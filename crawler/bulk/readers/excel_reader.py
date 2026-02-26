"""Excel source reader."""

from typing import Iterator, Dict, Any, List

from crawler.bulk.readers.base import SourceReader


class ExcelReader(SourceReader):
    """
    Reader for Excel files (.xlsx, .xls).
    
    Features:
    - Sheet selection
    - Streaming for large files
    - Type preservation
    """

    def __init__(
        self,
        file_path: str,
        sheet: int = 0,
        use_openpyxl: bool = True,
    ):
        super().__init__(file_path)
        self.sheet = sheet
        self.use_openpyxl = use_openpyxl
        self._wb = None

    def _open_workbook(self):
        """Open workbook if not already open."""
        if self._wb is None:
            from openpyxl import load_workbook
            self._wb = load_workbook(
                self.file_path,
                read_only=True,
                data_only=True,
            )

    def get_row_count(self) -> int:
        """Count rows in sheet."""
        self._open_workbook()
        ws = self._wb.worksheets[self.sheet]

        count = 0
        for _ in ws.iter_rows():
            count += 1

        # Subtract header
        return max(0, count - 1)

    def read_rows(self, batch_size: int = 100) -> Iterator[Dict[str, Any]]:
        """Read rows from Excel."""
        self._open_workbook()
        ws = self._wb.worksheets[self.sheet]

        # Get headers
        headers = None
        batch = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if headers is None:
                headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(row)]
                continue

            # Create dict from row
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = value

            batch.append(row_dict)

            if len(batch) >= batch_size:
                yield from batch
                batch = []

        if batch:
            yield from batch

    def get_columns(self) -> List[str]:
        """Get column names from sheet header."""
        self._open_workbook()
        ws = self._wb.worksheets[self.sheet]

        # Read first row as header
        headers = []
        for cell in next(ws.iter_rows(values_only=True)):
            headers.append(str(cell) if cell is not None else "")

        return headers

    def list_sheets(self) -> List[str]:
        """List sheet names."""
        self._open_workbook()
        return self._wb.sheetnames

    def validate(self) -> tuple[bool, list]:
        """Validate Excel file."""
        errors = []

        try:
            self._open_workbook()

            # Check sheet exists
            if self.sheet >= len(self._wb.worksheets):
                errors.append(f"Sheet index {self.sheet} not found")
                return False, errors

            # Check has data
            if self.get_row_count() == 0:
                errors.append("Sheet has no data rows")

        except Exception as e:
            errors.append(f"Failed to read Excel: {e}")

        return len(errors) == 0, errors

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self._wb:
            self._wb.close()
            self._wb = None
